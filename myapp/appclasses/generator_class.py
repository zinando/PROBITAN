import random
import datetime
import calendar
import openpyxl

my_list = [{'error': 'End Product', 'r1': 'unplanned stop', 'r2': 'No Powder Supply', 'r3':
    'Starvation from MSG failure', 'r4': 'MSG not producing (issue on MSG line)', 'comment': '',
            'freq': [2, 0, 3, 0, 1], 'mins': [60, 125, 45, 86, 246, 184], 'event_type': 'daily'},
           {'error': 'Power Fault', 'r1': 'unplanned stop', 'r2': 'IBEDC',
            'r3': 'Changeover from IBEDC to generator failure', 'r4': None, 'comment': '', 'freq': [1],
            'mins': [10, 15, 7, 12], 'event_type': 'weekly'},
           {'error': 'End Product', 'r1': 'unplanned stop', 'r2': 'No Powder Supply',
            'r3': 'Buggy not opened on time failure', 'r4': 'Contractor indiscipline', 'comment': '', 'freq': [5, 3, 4],
            'mins': [10, 5, 7, 12], 'event_type': 'daily'},
           {'error': 'Machine Stopped by Operator', 'r1': 'unplanned stop', 'r2': 'Polyfilm OOS',
            'r3': 'Ink stain failure', 'r4': 'Specify in comment', 'comment': '', 'freq': [1, 2, 4],
            'mins': [28, 35, 42], 'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'unplanned stop', 'r2': 'Contractor inefficiency',
            'r3': 'OTHER failure', 'r4': 'Specify in comment', 'comment': '', 'freq': [1, 3, 2], 'mins': [25, 38, 45],
            'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'unplanned stop', 'r2': 'Secondary Coding- Domino',
            'r3': 'Print head system failure', 'r4': 'Blockage in printhead', 'comment': '', 'freq': [1, 3, 2],
            'mins': [25, 38, 45], 'event_type': 'monthly'},
           {'error': 'Horizontal Jaw electric current limit.', 'r1': 'unplanned stop', 'r2': 'Control System',
            'r3': 'PLC failure', 'r4': 'Communication loss',
            'comment': 'Communication loss between PLC and extraction belt', 'freq': [4, 0, 3, 0],
            'mins': [125, 86, 73, 144], 'event_type': 'monthly'},
           {'error': 'Photoeye Fault', 'r1': 'unplanned stop', 'r2': 'Low Suction',
            'r3': 'Partial blockage on DCS line failure', 'r4': '', 'comment': '', 'freq': [1, 0], 'mins': [25, 38, 45],
            'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'unplanned stop', 'r2': 'Main Conveyor',
            'r3': 'OTHER failure', 'r4': 'Specify in comment',
            'comment': 'Product trapped between main and inclined conveyors', 'freq': [1, 3, 0, 2],
            'mins': [12, 25, 30], 'event_type': 'weekly'},
           {'error': '2D No Read', 'r1': 'unplanned stop', 'r2': 'Barcode reader', 'r3': '2D Camera failure',
            'r4': '2D NO READ (specify in comment)', 'comment': '', 'freq': [1, 0, 0, 1, 1], 'mins': [25, 38, 45],
            'event_type': 'monthly'},
           {'error': 'Temperature out of range', 'r1': 'unplanned stop', 'r2': 'Horizontal Sealing',
            'r3': 'Front heating element failure', 'r4': 'Broken heating element cable', 'comment': '',
            'freq': [1, 0, 0, 0], 'mins': [166, 105, 118, 76, 92, 147], 'event_type': 'monthly'},
           {'error': 'Temperature out of range', 'r1': 'unplanned stop', 'r2': 'Horizontal Sealing',
            'r3': 'Rear heating element failure', 'r4': 'Broken heating element cable', 'comment': '',
            'freq': [1, 0, 0, 0], 'mins': [125, 78, 145], 'event_type': 'monthly'},
           {'error': 'Temperature out of range', 'r1': 'unplanned stop', 'r2': 'Horizontal Sealing',
            'r3': 'Front thermocouple failure', 'r4': 'Broken thermocouple cable', 'comment': 'defective themocouple',
            'freq': [1, 0, 0, 0], 'mins': [35, 48, 55, 84], 'event_type': 'monthly'},
           {'error': 'Temperature out of range', 'r1': 'unplanned stop', 'r2': 'Horizontal Sealing',
            'r3': 'Rear thermocouple failure', 'r4': 'Broken thermocouple cable', 'comment': 'defective themocouple',
            'freq': [1, 0, 0, 0], 'mins': [35, 48, 55, 84], 'event_type': 'monthly'},
           {'error': 'Temperature out of range', 'r1': 'unplanned stop', 'r2': 'Vertical Sealing',
            'r3': 'Heating element failure', 'r4': 'Broken heating element cable', 'comment': '', 'freq': [1, 0, 0, 0],
            'mins': [35, 122, 90, 84], 'event_type': 'monthly'},
           {'error': 'Temperature out of range', 'r1': 'unplanned stop', 'r2': 'Vertical Sealing',
            'r3': 'Thermocouple failure', 'r4': 'Broken thermocouple cable', 'comment': 'defective themocouple',
            'freq': [1, 0, 1, 0], 'mins': [145, 148, 120, 84], 'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'unplanned stop', 'r2': 'Cutting and Perforation',
            'r3': 'Cylinder failure', 'r4': 'Damaged cylinder seal', 'comment': '', 'freq': [1, 0, 1, 0, 0],
            'mins': [65, 84, 155, 94], 'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'unplanned stop', 'r2': 'Primary Coding',
            'r3': 'Printhead failure', 'r4': 'Worn printhead', 'comment': '', 'freq': [1, 0, 1, 0],
            'mins': [125, 164, 108, 84], 'event_type': 'monthly'},
           {'error': 'Photoeye Fault', 'r1': 'unplanned stop', 'r2': 'Bag Forming', 'r3': 'Forming set failure',
            'r4': 'Contamination on formingset wall', 'comment': 'film jam on forming set', 'freq': [1, 0, 2, 0, 3, 0],
            'mins': [35, 25, 42, 69], 'event_type': 'daily'},
           {'error': 'Photoeye Fault', 'r1': 'unplanned stop', 'r2': 'Pulling', 'r3': 'Photocell failure',
            'r4': 'Photocell lost calibration', 'comment': '', 'freq': [1, 0, 1, 0, 0, 0], 'mins': [15, 23, 18, 29],
            'event_type': 'weekly'},
           {'error': 'Photoeye Fault', 'r1': 'Planned Stop', 'r2': 'Pack Material', 'r3': 'Manufacturer Splice',
            'r4': None, 'comment': '', 'freq': [4, 6, 3, 2], 'uptime': [75, 288, 125, 312, 277, 103, 81],
            'mins': [3, 2, 5, 7], 'event_type': 'minor'},
           {'error': 'Coder Fault', 'r1': 'Planned Stop', 'r2': 'Changeover', 'r3': 'Ribbon Changeover', 'r4': None,
            'comment': '', 'freq': [1, 0, 1], 'uptime': [75, 418, 366], 'mins': [5, 7, 3, 4], 'event_type': 'minor'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Vertical sealing',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [60, 45, 62, 59], 'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Spillage hopper',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [43, 39, 50, 45], 'event_type': 'weekly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on S-conveyor',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [115, 125, 112, 119], 'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Rear web system',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [65, 72, 50, 66], 'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Packing conveyor',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [105, 84, 71, 89], 'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Outfeed conveyor',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [124, 166, 192, 148], 'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Markem unit',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [65, 70, 72, 68], 'event_type': 'weekly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Main drive',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [64, 55, 70, 76], 'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Machine frame',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [85, 78, 64, 80], 'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Main conveyor',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [195, 142, 190], 'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Inclined conveyor',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [165, 127, 102, 130], 'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Imaje 9040',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [125, 133, 114, 106], 'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL',
            'r3': 'CIL on Horizontal sealing', 'r4': None, 'comment': '', 'freq': [1], 'mins': [155, 165, 160, 157],
            'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Front web system',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [136, 140, 124, 116, 128, 150], 'event_type': 'weekly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Forming set',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [105, 120, 126, 112, 90], 'event_type': 'weekly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Fischbein',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [160, 170, 165, 158], 'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Dosing unit',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [205, 222, 194], 'event_type': 'weekly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Triverter',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [34, 28, 27, 29], 'event_type': 'minor'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Triverter',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [30, 33, 34, 33], 'event_type': 'weekly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'CIL', 'r3': 'CIL on Triverter',
            'r4': None, 'comment': '', 'freq': [1], 'mins': [165, 164, 168, 162], 'event_type': 'monthly'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'RLS', 'r3': 'Shiftly RLS', 'r4': None,
            'comment': '', 'uptime': [726, 840, 923, 887], 'freq': [2], 'mins': [41, 45, 47, 38],
            'event_type': 'minor'},
           {'error': 'End Of Film', 'r1': 'Planned Stop', 'r2': 'Changeover', 'r3': 'Reel Changeover', 'r4': None,
            'comment': '', 'freq': [6, 4, 5, 5], 'uptime': [188, 215, 432, 294, 187, 313, 174], 'mins': [5, 7, 10, 6],
            'event_type': 'minor'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'Changeover', 'r3': 'Size Changeover',
            'r4': 'State SKU in comment', 'comment': '', 'freq': [1, 0, 2, 0], 'mins': [15, 17, 13, 11],
            'event_type': 'minor'},
           {'error': 'Machine Stopped by Operator', 'r1': 'Planned Stop', 'r2': 'Changeover', 'r3': 'Brand Changeover',
            'r4': 'State SKU in comment', 'comment': '', 'freq': [0, 1, 2, 0], 'mins': [25, 37, 33, 24],
            'event_type': 'minor'},
           {'error': 'Air Pressure Fault', 'r1': 'UTILITIES', 'r2': 'IBEDC', 'r3': 'Changeover from IBEDC to generator',
            'r4': None, 'comment': '', 'freq': [1, 0, 1, 1, 0, 1], 'mins': [15, 17, 13, 14], 'event_type': 'minor'},
           {'error': 'Air Pressure Fault', 'r1': 'UTILITIES', 'r2': 'Generator',
            'r3': 'Changeover from generator to IBEDC', 'r4': None, 'comment': '', 'freq': [1, 0, 1, 1, 0, 1],
            'mins': [15, 17, 13, 14], 'event_type': 'minor'}]

data = [{'machine': 'U',
         'events': [(1, 2022, 61), (2, 2022, 58), (3, 2022, 59), (4, 2022, 63), (5, 2022, 58), (6, 2022, 60),
                    (7, 2022, 72)]},
        {'machine': 'V',
         'events': [(1, 2022, 75), (2, 2022, 62), (3, 2022, 64), (4, 2022, 57), (5, 2022, 64), (6, 2022, 71),
                    (7, 2022, 60)]},
        {'machine': 'W',
         'events': [(1, 2022, 59), (2, 2022, 57), (3, 2022, 66), (4, 2022, 72), (5, 2022, 70), (6, 2022, 83),
                    (7, 2022, 64)]}]


class VolumeCalculator:
    """Class for calculating volumes based on time."""

    def calculate_volume_by_time(self, time_minutes, sku_grams, volume_per_time):
        """
        Calculate total volume in grams, kilograms, and metric tonnes based on time.

        Args:
        - time_minutes (int): Time in minutes
        - sku_grams (int): Grams per SKU
        - volume_per_time (int): Volume per unit of time

        Returns:
        - Tuple: Total volume in grams, kilograms, and metric tonnes
        """
        total_volume_grams = time_minutes * volume_per_time
        total_volume_kg = total_volume_grams / 1000
        total_volume_metric_tonnes = total_volume_kg / 1000

        return total_volume_grams, total_volume_kg, total_volume_metric_tonnes


class MachineMonthlyDowntimeLog:
    """Class for logging machine downtime events monthly."""

    def __init__(self, event_list, month, year, expected_pr, machine):
        """
        Initialize the MachineMonthlyDowntimeLog object.

        Args:
        - event_list (list): List of events
        - month (int): Month (1-12)
        - year (int): Year
        - expected_pr (float): Expected performance ratio
        """
        self.available_events = event_list
        self.month = month
        self.year = year
        self.expected_pr = expected_pr
        self.machine = machine
        self.event_log = []

        # Add machine attribute to each event in available_events
        for event in self.available_events:
            event['machine'] = machine

    def log_event(self):
        """
        Log a downtime event to the relevant method based on event type.

        This method calls each event log method without arguments in a try-except block.
        If successful, prints "Success" and returns self.event_log.
        If an error occurs, prints the error message, resets self.event_log to an empty list, and returns None.
        """
        output_folder = 'documents/output'

        try:
            self.log_daily_event()
            self.log_weekly_event()
            self.log_monthly_event()
            self.optimise_month_pr()
            print("Success")
            return True, self.event_log
        except Exception as e:
            print(f"Error: {e}")
            self.event_log = []  # Reset event log to an empty list on error
            return False, None

    def log_minor_event(self, event):
        """
        Log a minor event with downtime, uptime, and volume metrics.

        Args:
        - event (dict): Dictionary containing details of the event.

        Returns:
        - dict: Updated event dictionary with added downtime, uptime, and volume metrics.
        """
        # Assuming 'event' is a dictionary containing details of the event
        volume_calculator = VolumeCalculator()  # Instantiate VolumeCalculator if not already done

        # Generate start_time, end_time, and downtime for the event
        start_time, end_time, downtime = self.get_minor_event_valid_times()

        # Calculate downtime volume using VolumeCalculator
        downtime_volume_grams, _, _ = volume_calculator.calculate_volume_by_time(downtime, 85, 65)
        downtime_volume_kg = downtime_volume_grams / 1000

        # Calculate uptime using calculate_uptime method
        day_of_the_month = start_time.day  # Extracting the day from start_time
        uptime = self.calculate_uptime(day_of_the_month, start_time)

        # Calculate uptime volume using VolumeCalculator
        uptime_minutes = (end_time - start_time).total_seconds() / 60
        uptime_volume_grams, _, _ = volume_calculator.calculate_volume_by_time(uptime_minutes, 85, 65)
        uptime_volume_kg = uptime_volume_grams / 1000

        # Add start_time, end_time, downtime, uptime, and volume metrics to the event dictionary
        event['start_time'] = start_time
        event['end_time'] = end_time
        event['downtime'] = downtime
        event['uptime'] = uptime  # Add uptime to the event
        event['downtime_volume'] = downtime_volume_kg  # Add downtime_volume to the event
        event['uptime_volume'] = uptime_volume_kg  # Add uptime_volume to the event

        # Add the event to the event_log
        self.event_log.append(event)

        return event  # Return the updated event dictionary

    @staticmethod
    def get_num_days_in_week(year, week):
        """
        Get the number of days in a given week of a specific year.

        Args:
        - year (int): Year for which the week's days need to be counted.
        - week (int): Week number to calculate the days for.

        Returns:
        - int: Number of days in the specified week of the given year.
        """
        start_date = datetime.datetime.strptime(f'{year}-W{week}-1', "%Y-W%W-%w").date()
        end_date = start_date + datetime.timedelta(days=6)
        num_days_in_week = (end_date - start_date).days + 1
        return num_days_in_week

    @staticmethod
    def write_dict_to_excel(data, output_folder: str = 'documents/output'):
        """
        Writes data from a list of dictionaries into an Excel (.xlsx) file.

        Args:
        - data (list): A list of dictionaries containing the data to be written.
        - base_name (str): The base name for the output Excel file. Default is 'output'.

        Returns:
        - file_path (str): The file path including the filename where the Excel file is saved.
        """
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        file_name = f"output_{timestamp}.xlsx"
        file_path = f"{output_folder}/{file_name}"  # Update this path to your desired location

        wb = openpyxl.Workbook()
        sheet = wb.active

        # Write headers from the keys of the first dictionary
        headers = list(data[0].keys())
        headers_to_write = [header for header in headers if header != 'event_type'
                            and header != 'freq' and header != 'mins' and header != 'day_pr' and header != 'day_uptime']
        for col, header in enumerate(headers_to_write, start=1):
            sheet.cell(row=1, column=col).value = header

        # Write data into subsequent rows
        for row_idx, item in enumerate(data, start=2):
            for col_idx, key in enumerate(headers_to_write, start=1):
                value = item.get(key)
                # Check if the value is a list
                if isinstance(value, list):
                    # Convert the list to a string or handle it in a suitable way for Excel
                    formatted_value = ', '.join(
                        str(i) for i in value)  # Example: Convert list to a comma-separated string
                    sheet.cell(row=row_idx, column=col_idx).value = formatted_value
                else:
                    # For non-list values, write directly to the cell
                    sheet.cell(row=row_idx, column=col_idx).value = value

        # Save the workbook at the generated file path
        wb.save(file_path)
        return file_path

    @staticmethod
    def get_week_number(event_time):
        """
        Get the week number of a given event time.

        Args:
        - event_time (datetime): The timestamp of the event.

        Returns:
        - int: Week number of the provided event time.
        """
        event_date = event_time.date()
        year, week_num, _ = event_date.isocalendar()
        return week_num

    def get_valid_times(self, day):
        """
        Generate valid start and end times for downtime events for a given day.

        Args:
        - day (int): Day of the month

        Returns:
        - Tuple: (start_time, end_time, time_difference)
          - start_time (datetime): Start time for the downtime event
          - end_time (datetime): End time for the downtime event
          - time_difference (int): Time difference between start and end times in minutes
        """
        time_difference_list = [30, 60, 90, 120]  # Example list of time differences in minutes

        last_event_end_time = self.get_last_event_end_time_for_day(day)

        while True:
            start_hour = random.randint(0, 23)
            start_minute = random.randint(0, 59)
            start_time = datetime.datetime(self.year, self.month, day, start_hour, start_minute)

            time_difference = random.choice(time_difference_list)
            end_time = start_time + datetime.timedelta(minutes=time_difference)

            if start_time > last_event_end_time:  # Check if start_time is after the last event's end_time
                return start_time, end_time, time_difference

    def get_weekly_valid_times(self, week):
        """
        Generate valid start and end times for weekly downtime events for a given week.

        Args:
        - week (int): Week number

        Returns:
        - Tuple: (start_time, end_time, time_difference)
          - start_time (datetime): Start time for the weekly downtime event
          - end_time (datetime): End time for the weekly downtime event
          - time_difference (int): Time difference between start and end times in minutes
        """
        time_difference_list = [90, 120, 150, 180, 210, 240, 270, 300, 330, 360, 390, 420, 450, 480]

        # Determine the start and end dates for the given week
        # Adjust the calculation to ensure the day stays within the month's range
        start_date = datetime.datetime.strptime(f'{self.year}-W{week}-1', "%Y-W%W-%w").date()
        days_to_add = 6 - start_date.weekday()  # Days to add to reach the end of the week
        if start_date.month != (start_date + datetime.timedelta(days=days_to_add)).month:
            # If adding days exceeds the current month, adjust the start_date
            start_date = start_date.replace(day=1)  # Move to the first day of the next month
        end_date = start_date + datetime.timedelta(days=days_to_add)

        while True:
            # Randomly choose a day within the week
            random_day = random.randint(0, 6)
            start_hour = random.randint(0, 23)
            start_minute = random.randint(0, 59)

            # Calculate the start time within the selected day
            start_time = datetime.datetime(self.year, start_date.month, start_date.day + random_day, start_hour,
                                           start_minute)

            # Randomly choose the time difference between start and end times
            time_difference = random.choice(time_difference_list)
            end_time = start_time + datetime.timedelta(minutes=time_difference)

            # Check if the time difference is within the specified range
            if 90 <= time_difference <= 480:
                # Remove events within the time range from event_log
                self.event_log = [event for event in self.event_log if
                                  not (event['start_time'] >= start_time and event['end_time'] <= end_time)]

                # Adjust conflicting events in event_log (start_time < event['start_time'] < end_time)
                conflicting_events = [event for event in self.event_log if start_time <= event['start_time'] < end_time]
                for event in conflicting_events:
                    adjusted_start_time = max(end_time, event['end_time'] + datetime.timedelta(minutes=1))
                    if adjusted_start_time < event['end_time']:
                        event['start_time'] = adjusted_start_time
                    else:
                        # Efficiently remove conflicting event from event_log
                        self.event_log.remove(event)

                # Adjust conflicting events where event.end_time is between start_time and end_time
                conflicting_end_time_events = [event for event in self.event_log if
                                               end_time > event['end_time'] >= start_time > event[
                                                   'start_time']]
                for event in conflicting_end_time_events:
                    adjusted_end_time = min(start_time - datetime.timedelta(minutes=1), event['start_time'])
                    if adjusted_end_time > event['start_time']:
                        event['end_time'] = adjusted_end_time
                    else:
                        # Efficiently remove conflicting event from event_log
                        self.event_log.remove(event)

                return start_time, end_time, time_difference

    def get_monthly_valid_times(self):
        """
        Generate valid start and end times for monthly downtime events.

        Returns:
        - Tuple: (start_time, end_time, time_difference)
          - start_time (datetime): Start time for the monthly downtime event
          - end_time (datetime): End time for the monthly downtime event
          - time_difference (int): Time difference between start and end times in minutes
        """
        time_difference_list = [90, 120, 150, 180, 210, 240, 270, 300, 330, 360, 390, 420, 450, 480]

        # Get the number of days in the month
        num_days_in_month = calendar.monthrange(self.year, self.month)[1]

        # Randomly select a day within the month
        random_day = random.randint(1, num_days_in_month)

        # Randomly select start_time within the selected day
        start_hour = random.randint(0, 23)
        start_minute = random.randint(0, 59)
        start_time = datetime.datetime(self.year, self.month, random_day, start_hour, start_minute)

        # Randomly select time difference from the list
        time_difference = random.choice(time_difference_list)

        # Calculate end_time ensuring it's ahead of start_time and within the month
        end_time = start_time + datetime.timedelta(minutes=time_difference)
        if end_time.day != random_day:  # If end_time extends beyond the current day
            end_time = datetime.datetime(self.year, self.month, random_day, 23, 59)

        # Adjust end_time if it exceeds the last day of the month
        last_day_time = datetime.datetime(self.year, self.month, num_days_in_month, 23, 59)
        if end_time > last_day_time:
            end_time = last_day_time

        # Ensure end_time is ahead of start_time
        if end_time <= start_time:
            start_time = end_time - datetime.timedelta(minutes=time_difference)

        # Check if the time difference is within the specified range
        if 90 <= time_difference <= 480:
            # Remove events within the time range from event_log
            handled_events = [event for event in self.event_log if
                              not (event['start_time'] >= start_time and event['end_time'] <= end_time)]

            # Adjust conflicting events in event_log (start_time < event['start_time'] < end_time)
            conflicting_events = [event for event in handled_events if start_time <= event['start_time'] < end_time]
            modified_events = []
            for event in conflicting_events:
                adjusted_start_time = max(end_time, event['end_time'] + datetime.timedelta(minutes=1))
                if adjusted_start_time < event['end_time']:
                    event['start_time'] = adjusted_start_time
                    modified_events.append(event)

            non_conflicting_events = [event for event in handled_events if event not in conflicting_events]

            # Adjust conflicting events where event.end_time is between start_time and end_time
            conflicting_end_time_events = [event for event in non_conflicting_events if
                                           end_time > event['end_time'] >= start_time > event['start_time']]
            non_conflicting_end_time_events = [event for event in non_conflicting_events if
                                               event not in conflicting_events]

            for event in conflicting_end_time_events:
                adjusted_end_time = min(start_time - datetime.timedelta(minutes=1), event['start_time'])
                if adjusted_end_time > event['start_time']:
                    event['end_time'] = adjusted_end_time
                    modified_events.append(event)

            self.event_log = non_conflicting_end_time_events + modified_events

            return start_time, end_time, time_difference

    def get_minor_event_valid_times(self):
        """
        Generate valid start and end times for minor downtime events within a month.

        Returns:
        - Tuple: (start_time, end_time, time_difference)
          - start_time (datetime): Start time for the minor downtime event
          - end_time (datetime): End time for the minor downtime event
          - time_difference (int): Time difference between start and end times in minutes
        """
        # Define the range of time differences for minor events
        time_difference_list = list(range(2, 26))  # Generating a list from 2 to 25 (inclusive)

        # Get the number of days in the month
        num_days_in_month = calendar.monthrange(self.year, self.month)[1]

        # Randomly select a day within the month
        random_day = random.randint(1, num_days_in_month)

        # Randomly select start_time within the selected day
        start_hour = random.randint(0, 23)
        start_minute = random.randint(0, 59)
        start_time = datetime.datetime(self.year, self.month, random_day, start_hour, start_minute)

        # Randomly select time difference from the list
        time_difference = random.choice(time_difference_list)

        # Calculate end_time ensuring it's ahead of start_time and within the month
        end_time = start_time + datetime.timedelta(minutes=time_difference)
        if end_time.day != random_day:  # If end_time extends beyond the current day
            end_time = datetime.datetime(self.year, self.month, random_day, 23, 59)

        # Adjust end_time if it exceeds the last day of the month
        last_day_time = datetime.datetime(self.year, self.month, num_days_in_month, 23, 59)
        if end_time > last_day_time:
            end_time = last_day_time

        # Ensure end_time is ahead of start_time
        if end_time <= start_time:
            start_time = end_time - datetime.timedelta(minutes=time_difference)

        # Check if the time difference is within the specified range
        if 2 <= time_difference <= 25:
            # Remove events within the time range from event_log
            handled_events = [event for event in self.event_log if
                              not (event['start_time'] >= start_time and event['end_time'] <= end_time)]

            # Adjust conflicting events in event_log (start_time < event['start_time'] < end_time)
            conflicting_events = [event for event in handled_events if start_time <= event['start_time'] < end_time]
            modified_events = []
            for event in conflicting_events:
                adjusted_start_time = max(end_time, event['end_time'] + datetime.timedelta(minutes=1))
                if adjusted_start_time < event['end_time']:
                    event['start_time'] = adjusted_start_time
                    modified_events.append(event)

            non_conflicting_events = [event for event in handled_events if event not in conflicting_events]

            # Adjust conflicting events where event.end_time is between start_time and end_time
            conflicting_end_time_events = [event for event in non_conflicting_events if
                                           end_time > event['end_time'] >= start_time > event['start_time']]
            non_conflicting_end_time_events = [event for event in non_conflicting_events if
                                               event not in conflicting_events]

            for event in conflicting_end_time_events:
                adjusted_end_time = min(start_time - datetime.timedelta(minutes=1), event['start_time'])
                if adjusted_end_time > event['start_time']:
                    event['end_time'] = adjusted_end_time
                    modified_events.append(event)

            self.event_log = non_conflicting_end_time_events + modified_events

            return start_time, end_time, time_difference  # Return the generated times for minor events

    def get_last_event_end_time_for_day(self, day):
        """
        Retrieve the end time of the last event for a specific day.

        Args:
        - day (int): Day of the month to retrieve the last event's end time.

        Returns:
        - datetime: End time of the last event for the given day, or datetime.min if no events exist for that day.
        """
        events_for_day = [event for event in self.event_log if event.get('start_time').day == day]
        if events_for_day:
            last_event = max(events_for_day, key=lambda x: x['end_time'])
            return last_event['end_time'] if 'end_time' in last_event else datetime.datetime.min
        return datetime.datetime.min  # Return an early time if no events logged for the day

    def calculate_uptime(self, day, start_time):
        """
        Calculate uptime for a specific day based on the start time compared to the previous event's end time.

        Args:
        - day (int): Day of the month to calculate uptime for.
        - start_time (datetime): Start time of the event.

        Returns:
        - int: Uptime in minutes for the specified day. Returns 0 if no events are logged for the day.
        """
        preceding_events = [event for event in self.event_log if event.get('end_time') < start_time]
        if preceding_events:
            last_event = max(preceding_events, key=lambda x: x['end_time'])
            last_event_end_time = last_event['end_time']
            uptime = (start_time - last_event_end_time).seconds // 60
            return uptime
        return 0  # Return 0 if no events logged before the start_time on the day

    def calculate_day_uptime(self, day):
        """
        Calculate uptime and PR for a specific day.

        Args:
        - day (int): Day of the month to calculate uptime and PR for.

        Returns:
        - Tuple: (day_uptime, day_pr)
          - day_uptime (int): Uptime in minutes for the specified day.
          - day_pr (float): Percentage of uptime compared to the total minutes in the day.
        """
        total_minutes_in_day = 24 * 60
        day_events = [event for event in self.event_log if event.get('start_time').day == day]

        used_minutes = sum((event['end_time'] - event['start_time']).seconds // 60 for event in day_events)
        day_uptime = total_minutes_in_day - used_minutes
        day_pr = (day_uptime / total_minutes_in_day) * 100

        return day_uptime, day_pr

    def calculate_week_uptime_and_pr(self, week):
        """
        Calculate uptime and PR for a specific week.

        Args:
        - week (int): Week number to calculate uptime and PR for.

        Returns:
        - Tuple: (week_uptime, week_pr)
          - week_uptime (int): Uptime in minutes for the specified week.
          - week_pr (float): Percentage of uptime compared to the total minutes in the week.
        """
        week_events = [event for event in self.event_log if
                       MachineMonthlyDowntimeLog.get_week_number(event.get('start_time')) == week]
        week_downtime = sum(
            (event.get('end_time') - event.get('start_time')).total_seconds() / 60 for event in week_events)
        total_minutes_in_week = self.get_num_days_in_week(self.year, week) * 24 * 60
        week_uptime = total_minutes_in_week - week_downtime
        week_pr = (week_uptime / total_minutes_in_week) * 100 if total_minutes_in_week > 0 else 0
        return week_uptime, week_pr

    def calculate_month_uptime_and_pr(self):
        """
        Calculate uptime and PR for the entire month.

        Returns:
        - Tuple: (month_uptime, month_pr)
          - month_uptime (int): Total uptime in minutes for the entire month.
          - month_pr (float): Percentage of uptime compared to the total minutes in the month.
        """
        month_events = [event for event in self.event_log if event.get('start_time').month == self.month]
        month_downtime = sum(
            (event.get('end_time') - event.get('start_time')).total_seconds() / 60 for event in month_events)

        num_days_in_month = calendar.monthrange(self.year, self.month)[1]
        total_minutes_in_month = num_days_in_month * 24 * 60

        month_uptime = total_minutes_in_month - month_downtime
        month_pr = (month_uptime / total_minutes_in_month) * 100 if total_minutes_in_month > 0 else 0

        return month_uptime, month_pr

    def log_daily_event(self):
        """
        Log daily events based on availability and expected PR.

        Iterates through daily events, calculates uptime and volume metrics, and logs events
        into the event log if the daily uptime meets the expected PR.
        """
        print('loging daily events...')
        num_days = calendar.monthrange(self.year, self.month)[1]

        for day in range(1, num_days + 1):
            daily_events = [event for event in self.available_events if event.get('event_type') == 'daily']
            print('daily events count: ', len(daily_events))

            for event in daily_events:
                day_uptime, day_pr = self.calculate_day_uptime(day)

                if day_pr >= self.expected_pr:
                    start_time, end_time, downtime = self.get_valid_times(day)

                    uptime = self.calculate_uptime(day, start_time)

                    volume_calculator = VolumeCalculator()
                    uptime_volume_grams, _, _ = volume_calculator.calculate_volume_by_time(uptime, 85, 65)
                    uptime_volume_kg = uptime_volume_grams / 1000

                    downtime_volume_grams, _, _ = volume_calculator.calculate_volume_by_time(downtime, 85, 65)
                    downtime_volume_kg = downtime_volume_grams / 1000

                    event['start_time'] = start_time
                    event['end_time'] = end_time
                    event['downtime'] = downtime
                    event['uptime'] = uptime
                    event['uptime_volume'] = uptime_volume_kg
                    event['downtime_volume'] = downtime_volume_kg
                    event['day_uptime'] = day_uptime
                    event['day_pr'] = day_pr

                    self.event_log.append(event)
                    break
                else:
                    break

    def log_weekly_event(self):
        """
        Log weekly events based on availability and expected PR.

        Iterates through weekly events, calculates uptime and volume metrics, and logs events
        into the event log if the weekly uptime meets the expected PR.
        """
        print('loging weekly events...')
        num_weeks = int((calendar.monthrange(self.year, self.month)[1] - 1) / 7) + 1
        print('gotten week number')
        for week in range(1, num_weeks + 1):
            weekly_events = [event for event in self.available_events if
                             event.get('event_type') == 'weekly']
            print('weekly events count: ', len(weekly_events))
            for event in weekly_events:
                print('started logging week events...')
                # Calculate week_pr for the current week
                week_uptime, week_pr = self.calculate_week_uptime_and_pr(week)
                print('gotten pr')
                # Check week_pr against expected_pr
                if week_pr >= self.expected_pr:
                    start_time, end_time, downtime = self.get_weekly_valid_times(week)
                    print('gotten start time ...')
                    # Calculate day_of_the_week from start_time
                    day_of_the_week = start_time.weekday()
                    print('day of the week...')

                    # Calculate uptime using calculate_uptime method
                    uptime = self.calculate_uptime(day_of_the_week, start_time)
                    print('uptime gotten...')

                    # Add relevant attributes to the event
                    event['start_time'] = start_time
                    event['end_time'] = end_time
                    event['downtime'] = downtime
                    event['uptime'] = uptime
                    event['week_uptime'] = week_uptime
                    event['week_pr'] = week_pr

                    # Add event to event_log
                    self.event_log.append(event)
                    break
                else:
                    # Break the loop if week_pr is less than expected_pr
                    break

    def log_monthly_event(self):
        """
        Log monthly events based on availability and expected PR.

        Iterates through monthly events, calculates uptime and volume metrics, and logs events
        into the event log if the monthly uptime meets the expected PR.
        """
        print('loging monthly events...')
        monthly_events = [event for event in self.available_events if
                          event.get('event_type') == 'monthly']
        print('monthly events count: ', len(monthly_events))
        for event in monthly_events:
            while True:
                # Calculate month uptime and PR for each event
                month_uptime, month_pr = self.calculate_month_uptime_and_pr()

                # Check if month_pr meets expected_pr
                if month_pr >= self.expected_pr:
                    # Get start_time, end_time, and downtime using get_monthly_valid_times
                    start_time, end_time, downtime = self.get_monthly_valid_times()

                    # Calculate uptime using calculate_uptime method
                    day_of_the_month = start_time.day
                    uptime = self.calculate_uptime(day_of_the_month, start_time)

                    # Calculate uptime_volume and downtime_volume using VolumeCalculator
                    volume_calculator = VolumeCalculator()
                    uptime_volume_grams, _, _ = volume_calculator.calculate_volume_by_time(uptime, 85, 65)
                    uptime_volume_kg = uptime_volume_grams / 1000

                    downtime_volume_grams, _, _ = volume_calculator.calculate_volume_by_time(downtime, 85, 65)
                    downtime_volume_kg = downtime_volume_grams / 1000

                    # Assign month_pr and month_uptime to the event
                    event['month_uptime'] = month_uptime
                    event['month_pr'] = month_pr

                    # Perform operations and add event to event log
                    # Example processing based on uptime/pr goes here...
                    event['start_time'] = start_time
                    event['end_time'] = end_time
                    event['downtime'] = downtime
                    event['uptime'] = uptime  # Add uptime to the event
                    event['uptime_volume'] = uptime_volume_kg  # Add uptime_volume to the event
                    event['downtime_volume'] = downtime_volume_kg  # Add downtime_volume to the event
                    self.event_log.append(event)  # Add event to event log
                    break  # Break out of the while loop for the current event
                else:
                    # Break the loop for the current event if month_pr is less than expected_pr
                    break

    def optimise_month_pr(self):
        """
        Optimize monthly PR by adding minor events.

        Attempts to optimize the monthly PR by adding minor events to the available events list
        until the monthly uptime meets the expected PR.
        """
        print('optimising monthly events...')
        minor_events = [event for event in self.available_events if event.get('event_type') == 'minor']

        # Calculate month_uptime and month_pr using calculate_month_uptime_and_pr
        month_uptime, month_pr = self.calculate_month_uptime_and_pr()

        while month_pr > self.expected_pr:
            print('month pr: ', month_pr)
            # Pick a random event from minor_events
            random_event = random.choice(minor_events)

            # Call log_minor_event method with the random event
            self.log_minor_event(random_event)

            # Recalculate month_uptime and month_pr after adding the new event
            month_uptime, month_pr = self.calculate_month_uptime_and_pr()


def log_monthly_events_for_machine(machine_name, monthly_events_data):
    """
    Logs a list of monthly events data for a given machine using the MachineMonthlyDowntimeLog class.

    Args:
    - machine_name (str): The name of the machine for which events are being logged.
    - monthly_events_data (list): A list of dictionaries containing monthly events data.

    Returns:
    - list: The total events logged as a list.
    """
    total_events_logged = []

    for item in monthly_events_data:
        month_number, year, expected_pr = item
        event_logger = MachineMonthlyDowntimeLog(
            event_list=my_list,
            month=month_number,
            year=year,
            expected_pr=expected_pr,
            machine=machine_name
        )
        event_is_logged, events, = event_logger.log_event()
        print('total events for month {} is {}'.format(month_number, len(events)))
        if event_is_logged:
            total_events_logged.extend([x.copy() for x in events])

    return total_events_logged


def log_events_for_list_of_machines_data() -> list:
    """
    Logs monthly events for a list of machines based on provided data.

    Args:
    - data (list): A list of dictionaries containing machine and events data.
                   Each dictionary should have keys 'machine' and 'events'.

    Returns:
    - list: A list containing all the events logged across the machines.

    The function iterates through the provided data for each machine, logging monthly events using
    the 'log_monthly_events_for_machine' function. It accumulates the total events logged across all machines
    and returns a list containing these events.
    """
    total_events_logged2 = []

    for item in data:
        logger = log_monthly_events_for_machine(item['machine'], item['events'])
        if len(logger) > 0:
            total_events_logged2.extend([x.copy() for x in logger])
            print('Total events for machine {} is {}'.format(item['machine'], len(logger)))

    return total_events_logged2
